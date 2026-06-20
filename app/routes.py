from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

from .extensions import db
from .models import Fund, FundEntry

web = Blueprint("web", __name__)


def _parse_entry_form(form_data):
    date_str = (form_data.get("date") or "").strip()
    principal = float(form_data.get("principal") or 0)
    current_value = float(form_data.get("current_value") or 0)

    gains_raw = (form_data.get("gains") or "").strip()
    gains_percent_raw = (form_data.get("gains_percent") or "").strip()

    entry_date = datetime.strptime(date_str, "%Y-%m-%d").date()

    if gains_raw:
        gains = float(gains_raw)
    else:
        gains = current_value - principal

    if gains_percent_raw:
        gains_percent = float(gains_percent_raw)
    else:
        gains_percent = (gains / principal * 100) if principal else 0.0

    return entry_date, principal, current_value, gains, gains_percent


@web.get("/")
def dashboard():
    funds = Fund.query.order_by(Fund.name.asc()).all()
    return render_template("dashboard.html", funds=funds)


@web.post("/funds")
def add_fund():
    name = (request.form.get("name") or "").strip()

    if not name:
        flash("Fund name is required.", "error")
        return redirect(url_for("web.dashboard"))

    existing = Fund.query.filter(db.func.lower(Fund.name) == name.lower()).first()
    if existing:
        flash("Fund with this name already exists.", "error")
        return redirect(url_for("web.dashboard"))

    fund = Fund(name=name)
    db.session.add(fund)
    db.session.commit()

    flash("Fund created successfully.", "success")
    return redirect(url_for("web.fund_detail", fund_id=fund.id))


@web.get("/fund/<int:fund_id>")
def fund_detail(fund_id: int):
    fund = Fund.query.get_or_404(fund_id)
    entries = FundEntry.query.filter_by(fund_id=fund.id).order_by(FundEntry.date.asc()).all()

    chart_dates = [e.date.strftime("%d/%m/%Y") for e in entries]
    principal_data = [e.principal for e in entries]
    current_value_data = [e.current_value for e in entries]
    gains_data = [e.gains for e in entries]

    return render_template(
        "fund_detail.html",
        fund=fund,
        entries=entries,
        chart_dates=chart_dates,
        principal_data=principal_data,
        current_value_data=current_value_data,
        gains_data=gains_data,
    )


@web.post("/fund/<int:fund_id>/entries")
def add_entry(fund_id: int):
    fund = Fund.query.get_or_404(fund_id)

    try:
        entry_date, principal, current_value, gains, gains_percent = _parse_entry_form(
            request.form
        )
    except ValueError:
        flash("Invalid input. Check date and numeric values.", "error")
        return redirect(url_for("web.fund_detail", fund_id=fund.id))

    entry = FundEntry(
        fund_id=fund.id,
        date=entry_date,
        principal=principal,
        current_value=current_value,
        gains=gains,
        gains_percent=gains_percent,
    )

    db.session.add(entry)
    db.session.commit()

    flash("Fund data added successfully.", "success")
    return redirect(url_for("web.fund_detail", fund_id=fund.id))


@web.get("/fund/<int:fund_id>/entry/<int:entry_id>/edit")
def edit_entry_page(fund_id: int, entry_id: int):
    fund = Fund.query.get_or_404(fund_id)
    entry = FundEntry.query.filter_by(id=entry_id, fund_id=fund.id).first_or_404()
    return render_template("edit_entry.html", fund=fund, entry=entry)


@web.post("/fund/<int:fund_id>/entry/<int:entry_id>/edit")
def update_entry(fund_id: int, entry_id: int):
    fund = Fund.query.get_or_404(fund_id)
    entry = FundEntry.query.filter_by(id=entry_id, fund_id=fund.id).first_or_404()

    try:
        entry_date, principal, current_value, gains, gains_percent = _parse_entry_form(
            request.form
        )
    except ValueError:
        flash("Invalid input. Check date and numeric values.", "error")
        return redirect(url_for("web.edit_entry_page", fund_id=fund.id, entry_id=entry.id))

    # Always recalculate on edit so metrics and chart reflect updated values.
    gains = current_value - principal
    gains_percent = (gains / principal * 100) if principal else 0.0

    entry.date = entry_date
    entry.principal = principal
    entry.current_value = current_value
    entry.gains = gains
    entry.gains_percent = gains_percent

    db.session.commit()

    flash("Entry updated successfully.", "success")
    return redirect(url_for("web.fund_detail", fund_id=fund.id))


@web.post("/fund/<int:fund_id>/entry/<int:entry_id>/delete")
def delete_entry(fund_id: int, entry_id: int):
    fund = Fund.query.get_or_404(fund_id)
    entry = FundEntry.query.filter_by(id=entry_id, fund_id=fund.id).first_or_404()

    db.session.delete(entry)
    db.session.commit()

    flash("Entry deleted successfully.", "success")
    return redirect(url_for("web.fund_detail", fund_id=fund.id))


@web.get("/fund/<int:fund_id>/export")
def export_fund_excel(fund_id: int):
    fund = Fund.query.get_or_404(fund_id)
    entries = FundEntry.query.filter_by(fund_id=fund.id).order_by(FundEntry.date.asc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"

    sheet["A1"] = "Fund Name"
    sheet["B1"] = fund.name
    sheet["A2"] = "Generated At"
    sheet["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    latest = fund.latest_entry
    if latest:
        sheet["A4"] = "Latest Principal"
        sheet["B4"] = latest.principal
        sheet["A5"] = "Latest Current Value"
        sheet["B5"] = latest.current_value
        sheet["A6"] = "Latest Gains"
        sheet["B6"] = latest.gains
        sheet["A7"] = "Latest Gains %"
        sheet["B7"] = latest.gains_percent

    start_row = 10
    headers = ["Date", "Principal", "Current Value", "Gains", "Gains %"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=start_row, column=col, value=header)

    for i, entry in enumerate(entries, start=start_row + 1):
        sheet.cell(row=i, column=1, value=entry.date.strftime("%Y-%m-%d"))
        sheet.cell(row=i, column=2, value=entry.principal)
        sheet.cell(row=i, column=3, value=entry.current_value)
        sheet.cell(row=i, column=4, value=entry.gains)
        sheet.cell(row=i, column=5, value=entry.gains_percent)

    if entries:
        data_end_row = start_row + len(entries)

        chart = LineChart()
        chart.title = f"Performance - {fund.name}"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Date"

        values = Reference(sheet, min_col=2, max_col=4, min_row=start_row, max_row=data_end_row)
        categories = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=data_end_row)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 18
        sheet.add_chart(chart, "H10")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{fund.name.replace(' ', '_').lower()}_dashboard.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
